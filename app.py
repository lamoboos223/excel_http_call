import pandas as pd
import requests
from openpyxl import load_workbook


url_col = 'url'
file_path = 'Book1.xlsx'
sheet_name = "Sheet1"
result_col_letter = "B"
data = pd.read_excel(file_path, engine='openpyxl', usecols=[url_col])

wb = load_workbook(file_path)
ws = wb[sheet_name]

print("[INFO] Start Looping on excel workbook {} sheet {}".format(file_path, sheet_name))
for number, row in data.iterrows():
    for j, column in row.iteritems():
        url = column
        response_col = result_col_letter + str(number+2)
        
        # make http call
        print("[INFO] Calling the endpoint: /GET " + url)
        response = requests.get(url)
        response_text = response.text

        print("[INFO] Writing the http response to cell: " + response_col)

        # writing response to the proper cell in this example it was cell 'B' but you can change it to your preferred column
        ws[response_col] = response_text
        wb.save(file_path)

print("[INFO] Done!")
